from flask import Flask, request, Response, jsonify, send_file, redirect, url_for
from io import BytesIO
import uuid
import os
import logging
import traceback
import subprocess
import tempfile
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)

START_COLUMN = 8

# Global variables to store the latest analytic data and file
latest_analytic_date = None
latest_attendance_data = None  # {'attended': {}, 'not_attended': {}}
latest_file_stream = None
latest_week_display = None
latest_district_counts = None  # {'district': {'total': count, 'ages': {'age': count}}, '總計': total}
latest_main_district = None  # Main district name
all_attendance_data = []  # List of (date, {'attended': {}, 'not_attended': {}}, week_display) for all weeks

def chinese_to_int(chinese_num):
    """Convert Chinese numerals to Arabic integers."""
    numeral_map = {
        '一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
        '六': 6, '七': 7, '八': 8, '九': 9, '十': 10
    }
    return numeral_map.get(chinese_num, 0)

def convert_xls_to_xlsx(file_stream):
    """Convert .xls to .xlsx using soffice command."""
    logger.info("Converting .xls to .xlsx using soffice")
    file_stream.seek(0)
    file_content = file_stream.read()

    # Create a temporary file to store the .xls file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as temp_xls:
        temp_xls.write(file_content)
        temp_xls_path = temp_xls.name

    # Define the output .xlsx path
    temp_xlsx_path = temp_xls_path.replace('.xls', '.xlsx')

    try:
        # Run soffice command to convert .xls to .xlsx
        subprocess.run([
            'soffice',
            '--headless',
            '--convert-to',
            'xlsx',
            temp_xls_path,
            '--outdir',
            os.path.dirname(temp_xls_path)
        ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        logger.info(f"Successfully converted {temp_xls_path} to {temp_xlsx_path}")

        # Read the converted .xlsx file
        with open(temp_xlsx_path, 'rb') as temp_xlsx:
            output_stream = BytesIO(temp_xlsx.read())
        return output_stream

    except subprocess.CalledProcessError as e:
        logger.error(f"Failed to convert .xls to .xlsx: {e.stderr.decode()}")
        raise Exception("Failed to convert .xls to .xlsx")
    finally:
        # Clean up temporary files
        if os.path.exists(temp_xls_path):
            os.remove(temp_xls_path)
        if os.path.exists(temp_xlsx_path):
            os.remove(temp_xlsx_path)

def classify_attendance(sheet, week_col):
    global latest_main_district
    logger.debug(f"Classifying attendance for week column: {week_col}")
    attended = {}
    not_attended = {}
    district_counts = {}
    youth_above = {'年長', '中壯', '青壯', '青職'}
    age_categories = ['青職以上', '大專', '中學', '大學', '小學', '學齡前']
    max_row = sheet.max_row
    
    for row in range(3, max_row + 1):  # Start from row 3 (1-based index in openpyxl)
        main_district = str(sheet.cell(row, 1).value or "").strip()  # Column A
        sub_district = str(sheet.cell(row, 2).value or "").strip()  # Column B
        district = f"{main_district}{sub_district}"
        name = sheet.cell(row, 4).value  # Column D
        age = str(sheet.cell(row, 6).value or "").strip()  # Age in column F
        if not name or not district.startswith(main_district):
            continue
        # Set main district name from first valid row
        if latest_main_district is None and main_district:
            latest_main_district = main_district
            logger.debug(f"Set main district name to: {latest_main_district}")
        attendance = sheet.cell(row, week_col + 1).value  # week_col is 0-based, openpyxl is 1-based
        if attendance == 1:
            if district not in attended:
                attended[district] = []
            attended[district].append(name)
            if district not in district_counts:
                district_counts[district] = {'total': 0, 'ages': {age: 0 for age in age_categories}}
            district_counts[district]['total'] += 1
            # Map age to category
            effective_age = '青職以上' if age in youth_above or not age else age
            if effective_age not in age_categories:
                logger.warning(f"Unrecognized age '{age}' for {name} in {district}, defaulting to '青職以上'")
                effective_age = '青職以上'
            district_counts[district]['ages'][effective_age] += 1
        else:
            if district not in not_attended:
                not_attended[district] = []
            not_attended[district].append(name)
    total_attendance = sum(d['total'] for d in district_counts.values())
    district_counts['總計'] = total_attendance
    return attended, not_attended, district_counts

def write_summary(new_sheet, attended, not_attended):
    logger.debug(f"Writing summary with attended: {attended}, not_attended: {not_attended}")
    districts = sorted(set(attended.keys()).union(not_attended.keys()), key=lambda x: chinese_to_int(x[3:4]))
    row = 1  # 1-based index in openpyxl

    # Write headers with styles
    header_fill = PatternFill(start_color="107C10", end_color="107C10", fill_type="solid")
    subheader_fill = PatternFill(start_color="5DBB63", end_color="5DBB63", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for i, district in enumerate(districts):
        cell1 = new_sheet.cell(row, i * 2 + 1)
        cell2 = new_sheet.cell(row, i * 2 + 2)
        cell1.value = district
        cell2.value = district
        cell1.fill = header_fill
        cell2.fill = header_fill
        cell1.font = header_font
        cell2.font = header_font
        cell1.alignment = Alignment(horizontal='center')
        cell2.alignment = Alignment(horizontal='center')

        sub_cell1 = new_sheet.cell(row + 1, i * 2 + 1)
        sub_cell2 = new_sheet.cell(row + 1, i * 2 + 2)
        sub_cell1.value = "本週到會"
        sub_cell2.value = "未到會"
        sub_cell1.fill = subheader_fill
        sub_cell2.fill = subheader_fill
        sub_cell1.font = header_font
        sub_cell2.font = header_font
        sub_cell1.alignment = Alignment(horizontal='center')
        sub_cell2.alignment = Alignment(horizontal='center')

    max_len = max(max(len(attended.get(d, [])), len(not_attended.get(d, []))) for d in districts)
    for r in range(max_len):
        for i, district in enumerate(districts):
            attended_list = attended.get(district, [])
            not_attended_list = not_attended.get(district, [])
            if r < len(attended_list):
                new_sheet.cell(r + 3, i * 2 + 1).value = attended_list[r]
            if r < len(not_attended_list):
                new_sheet.cell(r + 3, i * 2 + 2).value = not_attended_list[r]

    logger.debug("Summary written successfully")

def process_excel(file_stream, file_extension):
    global latest_analytic_date, latest_attendance_data, latest_file_stream, latest_week_display, latest_district_counts, latest_main_district, all_attendance_data
    file_stream.seek(0)
    file_content = file_stream.read()
    buffered_stream = BytesIO(file_content)
    logger.info(f"Processing file with extension: {file_extension}, Size: {len(file_content)} bytes")

    # Convert .xls to .xlsx if necessary
    if file_extension == '.xls':
        logger.info("Detected .xls file, converting to .xlsx")
        buffered_stream = convert_xls_to_xlsx(file_stream)
        file_extension = '.xlsx'

    try:
        workbook = openpyxl.load_workbook(buffered_stream)
    except Exception as e:
        logger.error(f"Failed to load workbook: {str(e)}")
        raise

    input_sheet = workbook.active
    logger.debug(f"Loaded sheet: {input_sheet.title}, Rows: {input_sheet.max_row}, Columns: {input_sheet.max_column}")

    week_cols = []
    current_month = "2025年1月"
    for col in range(START_COLUMN, input_sheet.max_column + 1):
        month_header = str(input_sheet.cell(1, col + 1).value or "")
        week_header = str(input_sheet.cell(2, col + 1).value or "")
        if "2025年" in month_header:
            current_month = month_header.strip()
        if "週" in week_header:
            week_cols.append((col, week_header, current_month))

    logger.info(f"Detected week columns with months: {week_cols}")

    if not week_cols:
        logger.warning("No week columns detected; output will lack analytic sheets")

    # Clear previous data
    all_attendance_data = []
    latest_date = None
    latest_attended = None
    latest_not_attended = None
    latest_week = None
    latest_districts = None
    for col, week_name, month_prefix in week_cols:
        logger.info(f"Processing week: {week_name} in {month_prefix}")
        attended, not_attended, district_counts = classify_attendance(input_sheet, col)

        if not any(attended.values()):
            logger.info(f"No attendees for {week_name} in {month_prefix}, skipping sheet creation")
            continue

        new_sheet_name = f"{month_prefix}{week_name} 主日"
        week_str = week_name.replace("第", "").replace("週", "")
        week_num = chinese_to_int(week_str)
        month_num = int(month_prefix.split("年")[1].replace("月", ""))
        current_date = datetime(2025, month_num, min(week_num * 7, 28))
        
        # Store attendance data for this week
        all_attendance_data.append((current_date, {'attended': attended, 'not_attended': not_attended}, f"{month_prefix}{week_name}"))

        if latest_date is None or current_date > latest_date:
            latest_date = current_date
            latest_attended = attended
            latest_not_attended = not_attended
            latest_week = f"{month_prefix}{week_name}"
            latest_districts = district_counts

        if new_sheet_name in workbook.sheetnames:
            logger.error(f"Duplicate sheet name detected: {new_sheet_name}")
            raise ValueError(f"Sheet name '{new_sheet_name}' already exists")

        new_sheet = workbook.create_sheet(new_sheet_name)
        logger.debug(f"Created new sheet: {new_sheet_name}")
        write_summary(new_sheet, attended, not_attended)

    if latest_date:
        latest_analytic_date = latest_date.strftime("%Y年%m月%d日")
        latest_attendance_data = {'attended': latest_attended, 'not_attended': latest_not_attended}
        latest_week_display = latest_week
        latest_district_counts = latest_districts
        logger.debug(f"Updated latest_analytic_date to: {latest_analytic_date}, latest_week_display to: {latest_week_display}, latest_district_counts to: {latest_district_counts}")

    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    latest_file_stream = BytesIO(output_stream.read())
    logger.info("File processing completed successfully")
    return output_stream

@app.route('/')
def index():
    global latest_analytic_date, latest_attendance_data, latest_week_display, latest_district_counts, latest_main_district, all_attendance_data
    latest_date_display = latest_analytic_date if latest_analytic_date else "No analytics available yet"
    week_display = latest_week_display if latest_week_display else "No week data available yet"
    
    combined_table_html = ""
    if latest_attendance_data and latest_district_counts:
        districts = sorted(set(latest_attendance_data['attended'].keys()).union(latest_attendance_data['not_attended'].keys()), 
                          key=lambda x: chinese_to_int(x[3:4]))
        max_len = max(max(len(latest_attendance_data['attended'].get(d, [])), len(latest_attendance_data['not_attended'].get(d, []))) for d in districts)
        stats_districts = sorted([d for d in latest_district_counts.keys() if d != '總計'], key=lambda x: chinese_to_int(x[3:4]))
        age_categories = ['青職以上', '大專', '中學', '大學', '小學', '學齡前']
        
        # Find previous week's data
        previous_week_data = None
        if len(all_attendance_data) > 1:
            all_attendance_data.sort(key=lambda x: x[0])
            latest_date = all_attendance_data[-1][0]
            for date, data, week_name in reversed(all_attendance_data[:-1]):
                if date < latest_date:
                    previous_week_data = data
                    break

        # Prepare sorted lists with highlights
        sorted_attended = {}
        sorted_not_attended = {}
        for district in districts:
            attended_list = latest_attendance_data['attended'].get(district, [])
            not_attended_list = latest_attendance_data['not_attended'].get(district, [])
            
            attended_with_highlights = []
            not_attended_with_highlights = []
            if previous_week_data:
                prev_attended = previous_week_data['attended'].get(district, [])
                prev_not_attended = previous_week_data['not_attended'].get(district, [])
                
                for name in attended_list:
                    display_name = name[:4] if len(name) > 4 else name
                    highlight = 'highlight-green' if name in prev_not_attended else ''
                    attended_with_highlights.append((name, display_name, highlight))
                
                for name in not_attended_list:
                    display_name = name[:4] if len(name) > 4 else name
                    highlight = 'highlight-red' if name in prev_attended else ''
                    not_attended_with_highlights.append((name, display_name, highlight))
            
            else:
                attended_with_highlights = [(name, name[:4] if len(name) > 4 else name, '') for name in attended_list]
                not_attended_with_highlights = [(name, name[:4] if len(name) > 4 else name, '') for name in not_attended_list]
            
            # Sort: highlighted names come first
            attended_with_highlights.sort(key=lambda x: (x[2] == '', x[0]))
            not_attended_with_highlights.sort(key=lambda x: (x[2] == '', x[0]))
            
            sorted_attended[district] = attended_with_highlights
            sorted_not_attended[district] = not_attended_with_highlights
            max_len = max(max_len, max(len(attended_with_highlights), len(not_attended_with_highlights)))

        combined_table_html = """
        <div class="table-wrapper">
            <table class="excel-table">
        """

        # Precompute stats rows
        stats_rows = []
        row_index = 0
        for district in stats_districts:
            row_class = "even" if row_index % 2 == 0 else "odd"
            stats_rows.append((row_class, f'<td colspan="2" class="district-header">{district}</td>'))
            row_index += 1
            for age in age_categories:
                count = latest_district_counts[district]['ages'][age]
                row_class = "even" if row_index % 2 == 0 else "odd"
                stats_rows.append((row_class, f'<td class="sub-row" style="padding-left: 15px;">{age}</td><td class="sub-row">{count}</td>'))
                row_index += 1
            total = latest_district_counts[district]['total']
            row_class = "even" if row_index % 2 == 0 else "odd"
            stats_rows.append((row_class, f'<td class="sub-row" style="padding-left: 15px;">總計</td><td class="sub-row">{total}</td>'))
            row_index += 1

        main_district = latest_main_district if latest_main_district else "未知大區"
        overall_ages = {age: sum(latest_district_counts[d]['ages'][age] for d in stats_districts) for age in age_categories}
        total_attendance = latest_district_counts['總計']
        row_class = "even" if row_index % 2 == 0 else "odd"
        stats_rows.append((row_class, f'<td colspan="2" class="district-header">{main_district}</td>'))
        row_index += 1
        for age in age_categories:
            count = overall_ages[age]
            row_class = "even" if row_index % 2 == 0 else "odd"
            stats_rows.append((row_class, f'<td class="sub-row" style="padding-left: 15px;">{age}</td><td class="sub-row">{count}</td>'))
            row_index += 1
        row_class = "even" if row_index % 2 == 0 else "odd"
        stats_rows.append((row_class, f'<td class="sub-row" style="padding-left: 15px;">總計</td><td class="sub-row">{total_attendance}</td>'))

        stats_index = 0
        total_attendance_cols = len(districts) * 2

        combined_table_html += f'<tr class="title-row">'
        combined_table_html += f'<th colspan="{total_attendance_cols + 1}">{week_display}</th>'
        if stats_index < len(stats_rows):
            row_class, stats_cells = stats_rows[stats_index]
            combined_table_html += stats_cells
            stats_index += 1
        else:
            combined_table_html += '<td></td><td></td>'
        combined_table_html += '</tr>'

        combined_table_html += '<tr class="header">'
        for district in districts:
            combined_table_html += f'<th colspan="2">{district}</th>'
        combined_table_html += '<th class="separator"></th>'
        if stats_index < len(stats_rows):
            row_class, stats_cells = stats_rows[stats_index]
            combined_table_html += stats_cells
            stats_index += 1
        else:
            combined_table_html += '<td></td><td></td>'
        combined_table_html += '</tr>'

        combined_table_html += '<tr class="subheader">'
        for _ in districts:
            combined_table_html += '<th>本週到會</th><th>未到會</th>'
        combined_table_html += '<th class="separator"></th>'
        if stats_index < len(stats_rows):
            row_class, stats_cells = stats_rows[stats_index]
            combined_table_html += stats_cells
            stats_index += 1
        else:
            combined_table_html += '<td></td><td></td>'
        combined_table_html += '</tr>'

        for r in range(max_len):
            row_class = "even" if r % 2 == 0 else "odd"
            combined_table_html += f'<tr class="{row_class}">'
            for district in districts:
                attended_with_highlights = sorted_attended.get(district, [])
                not_attended_with_highlights = sorted_not_attended.get(district, [])
                attended_info = attended_with_highlights[r] if r < len(attended_with_highlights) else ('', '', '')
                not_attended_info = not_attended_with_highlights[r] if r < len(not_attended_with_highlights) else ('', '', '')
                attended_display = attended_info[1]
                not_attended_display = not_attended_info[1]
                attended_class = attended_info[2]
                not_attended_class = not_attended_info[2]

                combined_table_html += f'<td class="{attended_class}">{attended_display}</td><td class="{not_attended_class}">{not_attended_display}</td>'
            combined_table_html += '<td class="separator"></td>'
            if stats_index < len(stats_rows):
                row_class, stats_cells = stats_rows[stats_index]
                combined_table_html += stats_cells
                stats_index += 1
            else:
                combined_table_html += '<td></td><td></td>'
            combined_table_html += '</tr>'

        while stats_index < len(stats_rows):
            row_class, stats_cells = stats_rows[stats_index]
            combined_table_html += f'<tr class="{row_class}">'
            for _ in districts:
                combined_table_html += '<td></td><td></td>'
            combined_table_html += '<td class="separator"></td>'
            combined_table_html += stats_cells
            stats_index += 1
            combined_table_html += '</tr>'

        combined_table_html += "</table></div>"

    download_button = '<form action="/download" method="get"><input type="submit" value="Download Processed XLS" class="button"></form>' if latest_file_stream else ''
    
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            .table-wrapper {{
                overflow-x: auto;
                margin: 20px auto;
            }}
            .excel-table {{
                border-collapse: collapse;
                font-family: Arial, sans-serif;
                white-space: nowrap;
            }}
            .excel-table th, .excel-table td {{
                border: 1px solid #000;
                padding: 2px;
                text-align: left;
                vertical-align: top;
                min-width: 70px;
                line-height: 1.2;
            }}
            .excel-table .separator {{
                min-width: 10px;
                width: 10px;
            }}
            .excel-table .title-row th {{
                background-color: #005566;
                color: white;
                text-align: center;
                font-weight: bold;
                padding: 2px;
                line-height: 1.2;
            }}
            .excel-table .header th {{
                background-color: #107C10;
                color: white;
                padding: 2px;
                line-height: 1.2;
            }}
            .excel-table .subheader th {{
                background-color: #5DBB63;
                color: white;
                padding: 2px;
                line-height: 1.2;
            }}
            .excel-table tr.even {{
                background-color: #F3F2F1;
                color: black;
            }}
            .excel-table tr.odd {{
                background-color: #FFFFFF;
                color: black;
            }}
            .excel-table .sub-row {{
                background-color: #E1DFDD;
                font-size: 0.85em;
            }}
            .excel-table .district-header {{
                background-color: #107C10;
                color: white;
                text-align: center;
                font-weight: bold;
                padding: 2px;
                line-height: 1.2;
            }}
            .excel-table .highlight-green {{
                background-color: #90EE90;
            }}
            .excel-table .highlight-red {{
                background-color: #FFB6C1;
            }}
            .button {{
                background-color: #005566;
                color: white;
                padding: 8px 16px;
                border: none;
                cursor: pointer;
                margin-top: 10px;
            }}
            .button:hover {{
                background-color: #003f4c;
            }}
        </style>
    </head>
    <body>
        <h2>Upload Excel File for Analysis</h2>
        <p>Latest Analytic Date: {latest_date_display}</p>
        <form action="/upload" method="post" enctype="multipart/form-data">
            <input type="file" name="file" accept=".xls,.xlsx">
            <input type="submit" value="Upload and Analyze" class="button">
        </form>
        {download_button}
        <h3>Latest Attendance Data</h3>
        <div class="table-wrapper">
            {combined_table_html}
        </div>
    </body>
    </html>
    """

@app.route('/upload', methods=['POST'])
def upload_file():
    logger.info("Received upload request")
    if 'file' not in request.files:
        logger.error("No file uploaded")
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files['file']
    filename = file.filename.lower()
    logger.debug(f"Uploaded file: {filename}")
    if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
        logger.error("Invalid file format")
        return jsonify({"error": "Only .xls and .xlsx files are supported"}), 400
    
    file_extension = '.xls' if filename.endswith('.xls') else '.xlsx'
    
    try:
        process_excel(file.stream, file_extension)
        return redirect(url_for('index'))
    except Exception as e:
        logger.error(f"Processing error: {str(e)}")
        logger.debug(f"Full traceback: {traceback.format_exc()}")
        return jsonify({"error": f"Processing failed: {str(e)}"}), 500

@app.route('/download', methods=['GET'])
def download_file():
    global latest_file_stream
    if latest_file_stream is None:
        return jsonify({"error": "No processed file available"}), 404
    latest_file_stream.seek(0)
    return send_file(
        latest_file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"analyzed_{uuid.uuid4().hex}.xlsx"
    )

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    logger.info(f"Starting server on port {port}")
    app.run(debug=False, host='0.0.0.0', port=port)
