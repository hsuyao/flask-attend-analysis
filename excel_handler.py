import os
import subprocess
import tempfile
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from config import START_COLUMN, DB_TYPE
from utils import chinese_to_int, parse_district, parse_week_display
from database import get_six_month_averages, bulk_write, find_existing, get_all_latest_attendance_dates, get_week_attendance_count
import time
import logging
from pymongo import UpdateOne

logger = logging.getLogger(__name__)

def convert_xls_to_xlsx(file_stream):
    logger.info("Converting .xls to .xlsx using soffice")
    file_stream.seek(0)
    file_content = file_stream.read()

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as temp_xls:
        temp_xls.write(file_content)
        temp_xls_path = temp_xls.name

    temp_xlsx_path = temp_xls_path.replace('.xls', '.xlsx')

    try:
        result = subprocess.run([
            'soffice', '--headless', '--convert-to', 'xlsx',
            temp_xls_path, '--outdir', os.path.dirname(temp_xls_path)
        ], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        logger.info(f"Converted {temp_xls_path} to {temp_xlsx_path}")

        if not os.path.exists(temp_xlsx_path):
            logger.error("Converted .xlsx file not found")
            raise Exception("Conversion failed: output file not found")

        with open(temp_xlsx_path, 'rb') as temp_xlsx:
            output_stream = BytesIO(temp_xlsx.read())
        return output_stream

    except subprocess.CalledProcessError as e:
        logger.error(f"Failed to convert .xls to .xlsx: {e.stderr.decode()}")
        raise Exception(f"Conversion failed: {e.stderr.decode()}")
    finally:
        if os.path.exists(temp_xls_path):
            os.remove(temp_xls_path)
        if os.path.exists(temp_xlsx_path):
            os.remove(temp_xlsx_path)

def classify_attendance(sheet, week_col, week_display, placeholder_date, event_name):
    logger.info(f"Classifying attendance for column {week_col}, week_display: {week_display}, event_name: {event_name}")
    attended = {}
    not_attended = {}
    district_counts = {}
    main_district_counts = {}
    records = []
    youth_above = {'年長', '中壯', '青壯', '青職'}
    age_categories = ['青職以上', '大專', '中學', '大學', '小學', '學齡前']
    max_row = sheet.max_row
    main_district = None

    # First pass: collect all records and classify attendance
    temp_records = []
    for row in range(3, max_row + 1):
        main_district_value = str(sheet.cell(row, 1).value or "").strip()
        sub_district = str(sheet.cell(row, 2).value or "").strip()
        district = f"{main_district_value}{sub_district}"
        name = sheet.cell(row, 4).value
        age = str(sheet.cell(row, 6).value or "").strip()
        if not name or not district.startswith(main_district_value):
            continue
        if main_district is None and main_district_value:
            main_district = main_district_value

        attendance = sheet.cell(row, week_col + 1).value
        effective_age = '青職以上' if age in youth_above or not age else age
        if effective_age not in age_categories:
            effective_age = '青職以上'

        temp_records.append({
            "name": name,
            "date": placeholder_date.strftime('%Y-%m-%d'),
            "week_display": week_display,
            "attended": 1 if attendance == 1 else 0,
            "district": district,
            "age_group": effective_age,
            "event_name": event_name
        })

        if attendance == 1:
            attended.setdefault(district, []).append(name)
        else:
            not_attended.setdefault(district, []).append(name)

    # Second pass: filter records by district attendance
    district_attendance = {}
    for record in temp_records:
        district = record["district"]
        if district not in district_attendance:
            district_attendance[district] = sum(1 for r in temp_records if r["district"] == district and r["attended"] == 1)
        if district_attendance[district] == 0:
            logger.info(f"Skipping district {district} in {week_display} due to zero attendance")
            continue
        records.append(record)

        if record["attended"] == 1:
            district_counts.setdefault(district, {'total': 0, 'ages': {age: 0 for age in age_categories}})
            main_district_counts.setdefault(main_district, {'total': 0, 'ages': {age: 0 for age in age_categories}})
            district_counts[district]['total'] += 1
            main_district_counts[main_district]['total'] += 1
            district_counts[district]['ages'][record["age_group"]] += 1
            main_district_counts[main_district]['ages'][record["age_group"]] += 1

    total_attendance = sum(d['total'] for d in district_counts.values())
    district_counts['總計'] = total_attendance
    has_attendees = total_attendance > 0
    logger.info(f"Week {week_display} has attendees: {has_attendees}, total attendance: {total_attendance}")
    return attended, not_attended, district_counts, main_district, main_district_counts, records, has_attendees

def write_summary(new_sheet, attended, not_attended, week_display, previous_week_data=None):
    logger.info(f"Writing Excel summary for {week_display}")
    districts = sorted(set(attended.keys()).union(not_attended.keys()), key=lambda x: (chinese_to_int(x[0]), chinese_to_int(x[3:4])))
    row = 1

    header_fill = PatternFill(start_color="107C10", end_color="107C10", fill_type="solid")
    subheader_fill = PatternFill(start_color="5DBB63", end_color="5DBB63", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for i, district in enumerate(districts):
        for col, value in enumerate([district, district, district], i * 3 + 1):
            cell = new_sheet.cell(row, col)
            cell.value = value
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for col, value in enumerate(["本週到會", "未到會", "半年平均出席率"], i * 3 + 1):
            cell = new_sheet.cell(row + 1, col)
            cell.value = value
            cell.fill = subheader_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

    max_len = max(max(len(attended.get(d, [])), len(not_attended.get(d, []))) for d in districts)
    all_names = set()
    for district in districts:
        all_names.update(attended.get(district, []) + not_attended.get(district, []))
    avg_rates = get_six_month_averages(list(all_names), week_display)

    for i, district in enumerate(districts):
        attended_list = attended.get(district, [])
        not_attended_list = not_attended.get(district, [])
        combined_list = []
        if previous_week_data:
            prev_attended = previous_week_data['attended'].get(district, [])
            prev_not_attended = previous_week_data['not_attended'].get(district, [])
            combined_list.extend((name, True, name in prev_not_attended) for name in attended_list)
            combined_list.extend((name, False, name in prev_attended) for name in not_attended_list)
        else:
            combined_list.extend((name, True, False) for name in attended_list)
            combined_list.extend((name, False, False) for name in not_attended_list)

        combined_list.sort(key=lambda x: (-int(x[2]), x[0]))

        for r, (name, is_attended, has_highlight) in enumerate(combined_list[:max_len]):
            col_offset = i * 3 + 1
            if is_attended:
                cell = new_sheet.cell(r + 3, col_offset)
                cell.value = name
                if has_highlight:
                    cell.fill = green_fill
            else:
                cell = new_sheet.cell(r + 3, col_offset + 1)
                cell.value = name
                if has_highlight:
                    cell.fill = red_fill
            cell_rate = new_sheet.cell(r + 3, col_offset + 2)
            cell_rate.value = f"{avg_rates.get(name, 0.0):.2%}"

def process_excel(file_stream, file_extension, save_to_db=True):
    logger.info(f"Processing Excel file, extension: {file_extension}, save_to_db={save_to_db}")
    start_time = time.time()
    records_written = 0

    file_stream.seek(0)
    file_content = file_stream.read()
    buffered_stream = BytesIO(file_content)

    if file_extension == '.xls':
        buffered_stream = convert_xls_to_xlsx(file_stream)
        file_extension = '.xlsx'

    try:
        workbook = openpyxl.load_workbook(buffered_stream)
    except Exception as e:
        logger.error(f"Failed to load workbook: {str(e)}")
        raise

    input_sheet = workbook.active
    event_name = str(input_sheet.cell(1, 1).value or "未指定活動").strip()
    logger.info(f"Extracted event name: {event_name}")

    week_cols = []
    current_month = None
    for col in range(START_COLUMN, input_sheet.max_column + 1):
        month_header = str(input_sheet.cell(1, col + 1).value or "").strip()
        week_header = str(input_sheet.cell(2, col + 1).value or "").strip()
        if "年" in month_header and "月" in month_header:
            current_month = month_header
        if "週" in week_header and current_month:
            week_cols.append((col, week_header, current_month))

    logger.info(f"Detected week columns: {week_cols}")
    if not week_cols:
        logger.warning("No week columns detected")
        return {
            'latest_analytic_date': None,
            'latest_attendance_data': None,
            'latest_week_display': None,
            'latest_district_counts': None,
            'latest_main_district': None,
            'latest_main_district_counts': None,
            'all_attendance_data': [],
            'event_name': event_name,
            'records_written': 0,
            'week_avg_rates': {}  # Added to store per-week attendance rates
        }

    all_attendance_data = []
    latest_week = None
    latest_attended = None
    latest_not_attended = None
    latest_districts = None
    latest_main_district = None
    latest_main_district_counts = None
    all_records = []
    all_names = set()
    week_avg_rates = {}  # Store attendance rates per week

    for col, week_name, month_prefix in week_cols:
        logger.info(f"Processing week: {week_name} in {month_prefix}")
        year = int(month_prefix.split("年")[0])
        month_num = int(month_prefix.split("年")[1].replace("月", ""))
        placeholder_date = datetime(year, month_num, 1)
        week_display = f"{month_prefix}{week_name}"

        attended, not_attended, district_counts, main_district, main_district_counts, records, has_attendees = classify_attendance(
            input_sheet, col, week_display, placeholder_date, event_name
        )
        all_records.extend(records)
        if main_district and not latest_main_district:
            latest_main_district = main_district

        week_names = set()
        for district in attended:
            week_names.update(attended[district])
        for district in not_attended:
            week_names.update(not_attended[district])
        all_names.update(week_names)

        # Calculate attendance rates for this week
        if week_names:
            week_avg_rates[week_display] = get_six_month_averages(list(week_names), week_display)
            logger.debug(f"Calculated attendance rates for {week_display}: {week_avg_rates[week_display]}")

        all_attendance_data.append((placeholder_date, {'attended': attended, 'not_attended': not_attended}, week_display, event_name))

        if has_attendees or not latest_attended:
            latest_attended = attended
            latest_not_attended = not_attended
            latest_week = week_display
            latest_districts = district_counts
            latest_main_district_counts = main_district_counts

    all_attendance_data.sort(key=lambda x: parse_week_display(x[2]))

    if save_to_db and all_records and all_names:
        week_displays = list(set(r["week_display"] for r in all_records))
        logger.debug(f"Checking existing records for {len(all_names)} names and {len(week_displays)} week_displays")
        existing_keys = find_existing(list(all_names), week_displays, event_names=[event_name])
        existing_cache = set((name, week, evt) for name, week, evt in existing_keys)

        operations = []
        for week_data in all_attendance_data:
            _, data, week_display, evt_name = week_data
            has_attendees = any(data['attended'].values())
            if not has_attendees:
                logger.info(f"Skipping week {week_display} due to no attendees")
                continue

            # Check attendance count in database for each district
            week_records = [r for r in all_records if r["week_display"] == week_display and r["event_name"] == evt_name]
            districts = set(r["district"] for r in week_records)
            for district in districts:
                db_attendance_count = get_week_attendance_count(week_display, district, evt_name)
                logger.debug(f"DB attendance count for {week_display}, {district}: {db_attendance_count}")
                district_records = [r for r in week_records if r["district"] == district]
                
                for record in district_records:
                    key = (record["name"], week_display, evt_name)
                    operations.append(UpdateOne(
                        {
                            "name": record["name"],
                            "week_display": week_display,
                            "event_name": evt_name
                        },
                        {"$set": record},
                        upsert=True
                    ))
                    logger.debug(
                        f"Added UpdateOne for {record['name']} in {week_display}, {district} "
                        f"(exists: {key in existing_cache})"
                    )

        logger.info(f"Prepared {len(operations)} bulk operations")
        if operations:
            try:
                bulk_result = bulk_write(operations)
                if bulk_result:
                    records_written += bulk_result.inserted_count + bulk_result.modified_count
                    logger.info(f"Bulk write result: {bulk_result.inserted_count} inserted, "
                               f"{bulk_result.modified_count} modified, {bulk_result.matched_count} matched")
                else:
                    logger.warning("No records written in bulk_write")
            except Exception as e:
                logger.error(f"Failed to write records: {str(e)}")
                raise

        if latest_week and latest_attended:
            supplement_start = time.time()
            existing_keys = find_existing(list(all_names), [latest_week], event_names=[event_name])
            existing_cache.update((name, week, evt) for name, week, evt in existing_keys)

            missing_records = []
            for name in all_names:
                if (name, latest_week, event_name) not in existing_cache:
                    district = next((d for d in latest_attended if name in latest_attended.get(d, [])), None) or \
                              next((d for d in latest_not_attended if name in latest_not_attended.get(d, [])), None) or "未知區"
                    db_attendance_count = get_week_attendance_count(latest_week, district, event_name)
                    if db_attendance_count == 0:
                        missing_records.append({
                            "name": name,
                            "date": datetime.now().strftime("%Y-%m-%d"),
                            "week_display": latest_week,
                            "attended": 0,
                            "district": district,
                            "age_group": "未知",
                            "event_name": event_name
                        })
                        logger.debug(f"Added missing record for {name} in {latest_week}, {district} "
                                    f"(DB attendance: {db_attendance_count})")

            if missing_records:
                logger.info(f"Writing {len(missing_records)} missing records")
                try:
                    bulk_result = bulk_write([UpdateOne(
                        {
                            "name": r["name"],
                            "week_display": r["week_display"],
                            "event_name": r["event_name"]
                        },
                        {"$set": r},
                        upsert=True
                    ) for r in missing_records])
                    if bulk_result:
                        records_written += bulk_result.inserted_count + bulk_result.modified_count
                        logger.info(f"Wrote {bulk_result.inserted_count} missing records, "
                                  f"{bulk_result.modified_count} modified for {latest_week}")
                    else:
                        logger.warning("No missing records written in bulk_write")
                except Exception as e:
                    logger.error(f"Failed to write missing records: {str(e)}")
                    raise

            supplement_elapsed = time.time() - supplement_start
            logger.info(f"Missing records supplement completed in {supplement_elapsed:.2f}s")

    if not all_attendance_data:
        logger.warning("No valid attendance data processed")
        return {
            'latest_analytic_date': None,
            'latest_attendance_data': None,
            'latest_week_display': None,
            'latest_district_counts': None,
            'latest_main_district': None,
            'latest_main_district_counts': None,
            'all_attendance_data': [],
            'event_name': event_name,
            'records_written': 0,
            'week_avg_rates': {}
        }

    total_elapsed = time.time() - start_time
    logger.info(f"Excel processing completed in {total_elapsed:.2f}s, {records_written} records written")
    return {
        'latest_analytic_date': datetime.now().strftime("%Y年%m月%d日"),
        'latest_attendance_data': {'attended': latest_attended, 'not_attended': latest_not_attended} if latest_attended else None,
        'latest_week_display': latest_week,
        'latest_district_counts': latest_districts,
        'latest_main_district': latest_main_district,
        'latest_main_district_counts': latest_main_district_counts,
        'all_attendance_data': all_attendance_data,
        'event_name': event_name,
        'records_written': records_written,
        'week_avg_rates': week_avg_rates
    }

def generate_excel(all_attendance_data):
    logger.info("Generating Excel file")
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    all_attendance_data.sort(key=lambda x: parse_week_display(x[2]))

    for date, data, week_name, event_name in all_attendance_data:
        new_sheet_name = f"{week_name} 主日"

        if new_sheet_name in workbook.sheetnames:
            logger.error(f"Duplicate sheet name: {new_sheet_name}")
            raise ValueError(f"Sheet name '{new_sheet_name}' exists")

        new_sheet = workbook.create_sheet(new_sheet_name)
        previous_week_data = None
        current_week_idx = next(idx for idx, (_, _, w, _) in enumerate(all_attendance_data) if w == week_name)
        if current_week_idx > 0:
            previous_week_data = all_attendance_data[current_week_idx - 1][1]

        write_summary(new_sheet, data['attended'], data['not_attended'], week_name, previous_week_data)

    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    return output_stream